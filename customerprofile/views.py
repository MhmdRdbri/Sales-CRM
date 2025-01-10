from rest_framework import generics
from rest_framework.views import APIView
from .models import CustomerProfile
from .serializers import *
from rest_framework.permissions import *
from rest_framework.exceptions import PermissionDenied
from factors.models import Factors
from django.db.models import Avg, Sum
from rest_framework.parsers import MultiPartParser
from openpyxl import load_workbook
from rest_framework.response import Response

class CustomerProfileListCreateView(generics.ListCreateAPIView):
    queryset = CustomerProfile.objects.all()
    serializer_class = CustomerProfileSerializer
    permission_classes = [IsAuthenticated]

class CustomerProfileRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CustomerProfile.objects.all()
    serializer_class = CustomerProfileSerializer
    permission_classes = [IsAuthenticated]

    
    def destroy(self, request, *args, **kwargs):
        print(f"User's work_position: {getattr(request.user.profile, 'work_position', 'No Profile')}")

        if hasattr(request.user, 'profile') and request.user.profile.work_position == 'regular':
            raise PermissionDenied("Employees are not allowed to delete customer profiles.")
        
        return super().destroy(request, *args, **kwargs)

class UploadCustomerProfilesView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        # Ensure a file is provided in the request
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file uploaded'}, status=400)

        # Load the uploaded Excel file
        try:
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active
        except Exception as e:
            return Response({'error': f'Error reading file: {str(e)}'}, status=400)

        # Process each row and create CustomerProfile entries
        created_profiles = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):  # Skip the header row
            full_name = row[0]
            phone_number = row[1]

            if not full_name or not phone_number:
                continue  # Skip rows with missing data

            # Create or update the customer profile
            profile, created = CustomerProfile.objects.update_or_create(
                phone_number=phone_number,
                defaults={'full_name': full_name}
            )
            created_profiles.append(profile.full_name)

        return Response({'message': 'Profiles processed successfully', 'profiles': created_profiles})